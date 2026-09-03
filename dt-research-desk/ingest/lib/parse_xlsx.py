#!/usr/bin/env python3
"""corpus/xlsx/{SP1500,Thematic}_<YYYY-MM-DD>.xlsx -> data/rankings.json"""
import openpyxl, os, json, glob, re
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC, OUT = os.path.join(ROOT, "corpus", "xlsx"), os.path.join(ROOT, "data", "rankings.json")

def rows(ws, hdr_row=3):
    hdr, out = None, []
    for i, r in enumerate(ws.iter_rows(values_only=True), 1):
        if i < hdr_row:
            continue
        if i == hdr_row:
            hdr = [str(c).replace("\n", " ").strip() if c is not None else "" for c in r]
            continue
        if all(c is None for c in r):
            continue
        out.append(dict(zip(hdr, r)))
    return out

def ranked(ws):
    return [r for r in rows(ws) if str(r.get("Rank", "")).strip().isdigit()]

DB = {"sp1500": {}, "thematic": {}}
SEC = ["Rank","Sector","# Stocks","Avg Score","Bull/Bull","Bull/Bear","Bear/Bull","Bear/Bear","Top Name"]
SUB = ["Rank","Sub-Industry","Sector","# Stocks","Avg Score","Bull/Bull","Bull/Bear","Bear/Bull","Bear/Bear"]
STK = ["Rank","Ticker","Company Name","Index","Sector","Sub-Industry","LT Sig","LT Days","ST Sig","ST Days"]

for f in sorted(glob.glob(os.path.join(SRC, "SP1500_*.xlsx"))):
    date = re.search(r"(\d{4}-\d{2}-\d{2})", f).group(1)
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    allr = ranked(wb["All Sectors"])
    cnt = {"BB": 0, "BBear": 0, "BearB": 0, "BearBear": 0}
    for r in allr:
        lt, st = str(r.get("LT Sig", "")), str(r.get("ST Sig", ""))
        k = "BB" if lt == "Bullish" and st == "Bullish" else \
            "BBear" if lt == "Bullish" else "BearB" if st == "Bullish" else "BearBear"
        cnt[k] += 1
    DB["sp1500"][date] = {
        "sectors": [{k: r.get(k) for k in SEC} for r in ranked(wb["Sector Summary"])],
        "subind":  [{k: r.get(k) for k in SUB} for r in ranked(wb["Sub-Industry Rankings"])][:25],
        "top":     [{k: r.get(k) for k in STK} for r in allr[:50]],
        "universe": cnt, "n": len(allr)}
    wb.close()
    print("SP1500 %s  stocks=%d  dual-bull=%.1f%%" % (date, len(allr), 100 * cnt["BB"] / max(len(allr), 1)))

THM = ["Rank","Theme","ETF","# Stocks","Avg Score","Bull/Bull","Bull/Bear","Bear/Bull","Bear/Bear"]
TSK = ["Rank","Ticker","Company Name","ETF","Thematic Category","LT Sig","ST Sig"]
for f in sorted(glob.glob(os.path.join(SRC, "Thematic_*.xlsx"))):
    date = re.search(r"(\d{4}-\d{2}-\d{2})", f).group(1)
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    allr = ranked(wb["Thematic Rankings"])
    DB["thematic"][date] = {
        "themes": [{k: r.get(k) for k in THM} for r in ranked(wb["Theme Rankings"])],
        "top": [{k: r.get(k) for k in TSK} for r in allr[:20]], "n": len(allr)}
    wb.close()
    print("Thematic %s  %d stocks" % (date, len(allr)))

json.dump(DB, open(OUT, "w"), indent=1, default=str)
print("wrote", OUT)
